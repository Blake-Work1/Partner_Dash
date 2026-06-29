#!/usr/bin/env python3
"""
Partner Performance Cube Data Aggregation
Reads SFDC export, applies partner/team lookups, pre-computes all filter combinations
Outputs JSON for embedding in HTML dashboard
"""

import pandas as pd
import json
from datetime import datetime
from collections import defaultdict
import openpyxl

# ============================================================================
# LOAD LOOKUPS
# ============================================================================

def load_partner_lookups():
    """Extract partner classification and tier mapping"""
    wb = openpyxl.load_workbook('/mnt/user-data/uploads/JA_Look_UP.xlsx')
    ws = wb['Partner Look UP']
    
    partners = {}
    seen = set()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        account_name = row[0]
        sfdc_id = row[6]
        partner_class = row[4]
        top_level = row[3]
        geo = row[5]
        
        if sfdc_id and account_name and partner_class:
            key = str(sfdc_id).strip()
            if key not in seen and key != 'None':
                partners[key] = {
                    'name': str(account_name),
                    'top_level': str(top_level) if top_level else str(account_name),
                    'tier': str(partner_class),
                    'geo': str(geo) if geo else None
                }
                seen.add(key)
    
    return partners

def load_team_lookups():
    """Extract team mapping (booking team -> final team)"""
    wb = openpyxl.load_workbook('/mnt/user-data/uploads/JA_Look_UP.xlsx')
    ws = wb['Sheet1']
    
    teams = {}
    seen = set()
    
    for row in ws.iter_rows(min_row=1, values_only=True):
        booking_team = row[0]
        final_team = row[1]
        
        if booking_team and final_team and booking_team not in seen:
            bt = str(booking_team).strip()
            ft = str(final_team).strip()
            if bt != 'None' and ft != 'None':
                teams[bt] = ft
                seen.add(booking_team)
    
    return teams

# ============================================================================
# LOAD & CLEAN DATA
# ============================================================================

def load_data(filepath):
    """Load SQL export, clean dates, handle NULLs"""
    df = pd.read_excel(filepath, sheet_name='Sheet1')
    
    # Clean column names
    df.columns = df.columns.str.strip()
    
    # Convert Close_Date to datetime
    df['Close_Date'] = pd.to_datetime(df['Close_Date'], errors='coerce')
    
    # Filter out future dates (>= 2099)
    df = df[df['Close_Date'].dt.year < 2099]
    
    # Fill NULLs
    df['Territory'] = df['Territory'].fillna('Unknown')
    df['Source'] = df['Source'].fillna('Unattributed')
    df['Sourcing Partner'] = df['Sourcing Partner'].fillna('None')
    df['Influence Partner'] = df['Influence Partner'].fillna('None')
    df['Service Delivery Partner'] = df['Service Delivery Partner'].fillna('None')
    df['Tier'] = df['Tier'].fillna('Standard')
    
    # Replace 'NULL' string with None
    for col in df.columns:
        df[col] = df[col].replace('NULL', None)
    
    print(f"Loaded {len(df)} records after filtering future dates")
    return df

# ============================================================================
# ENRICHMENT
# ============================================================================

def enrich_data(df, partner_lookup, team_lookup):
    """Add partner tier/top_level, map territories"""
    
    # Enrich sourcing partners
    df['sourcing_tier'] = df['Sourcing Partner'].apply(
        lambda x: partner_lookup.get(x, {}).get('tier', 'Unclassified') if x and x != 'None' else None
    )
    df['sourcing_top_level'] = df['Sourcing Partner'].apply(
        lambda x: partner_lookup.get(x, {}).get('top_level', x) if x and x != 'None' else None
    )
    
    # Enrich influence partners
    df['influence_tier'] = df['Influence Partner'].apply(
        lambda x: partner_lookup.get(x, {}).get('tier', 'Unclassified') if x and x != 'None' else None
    )
    df['influence_top_level'] = df['Influence Partner'].apply(
        lambda x: partner_lookup.get(x, {}).get('top_level', x) if x and x != 'None' else None
    )
    
    # Enrich service partners
    df['service_tier'] = df['Service Delivery Partner'].apply(
        lambda x: partner_lookup.get(x, {}).get('tier', 'Unclassified') if x and x != 'None' else None
    )
    df['service_top_level'] = df['Service Delivery Partner'].apply(
        lambda x: partner_lookup.get(x, {}).get('top_level', x) if x and x != 'None' else None
    )
    
    # Map territories
    df['Territory_Mapped'] = df['Territory'].apply(
        lambda x: team_lookup.get(x, x) if x else 'Unknown'
    )
    
    # Determine if deal was won
    df['Won'] = df['Stage'].isin(['Closed Won', 'Stage 5 - Closed Won', '6 - Closed/Pending']).astype(int)
    
    # Extract month for grouping
    df['Month'] = df['Close_Date'].dt.to_period('M')
    df['Quarter'] = df['Close_Date'].dt.to_period('Q')
    
    return df

# ============================================================================
# AGGREGATION
# ============================================================================

def aggregate_partners(df, partner_col, tier_col, top_level_col, contribution_type):
    """
    Aggregate metrics by all filter combinations for one partner type
    Returns dict: {(geo, territory, product, source, deal_type, month): {partner: {metrics}}}
    """
    
    # Filter to rows with actual partners
    df_partners = df[df[partner_col] != 'None'].copy()
    
    if len(df_partners) == 0:
        return {}
    
    aggregates = defaultdict(lambda: defaultdict(lambda: {
        'nacv': 0, 'count': 0, 'won': 0, 'lost': 0, 
        'contribution_type': contribution_type
    }))
    
    for _, row in df_partners.iterrows():
        key = (
            row['Geo'],
            row['Territory_Mapped'],
            row['Product'],
            row['Source'],
            row['Deal Type'],
            str(row['Month']) if pd.notna(row['Month']) else 'Unknown'
        )
        
        partner_key = row[partner_col]
        
        agg = aggregates[key][partner_key]
        agg['nacv'] += row['Product NACV'] if pd.notna(row['Product NACV']) else 0
        agg['count'] += 1
        agg['won'] += row['Won']
        agg['lost'] += (1 - row['Won'])
        agg['tier'] = row[tier_col] if pd.notna(row[tier_col]) else 'Unclassified'
        agg['top_level'] = row[top_level_col] if pd.notna(row[top_level_col]) else partner_key
    
    return aggregates

def build_dashboard_data(df):
    """Pre-compute all aggregations for dashboard"""
    
    sourcing = aggregate_partners(df, 'Sourcing Partner', 'sourcing_tier', 'sourcing_top_level', 'Originated')
    influence = aggregate_partners(df, 'Influence Partner', 'influence_tier', 'influence_top_level', 'Influenced')
    service = aggregate_partners(df, 'Service Delivery Partner', 'service_tier', 'service_top_level', 'Resale/MSP')
    
    # Merge all partner types
    all_partners = defaultdict(lambda: defaultdict(dict))
    
    for key, partners in sourcing.items():
        for partner_name, metrics in partners.items():
            all_partners[key][partner_name].update(metrics)
    
    for key, partners in influence.items():
        for partner_name, metrics in partners.items():
            if partner_name in all_partners[key]:
                all_partners[key][partner_name]['influence_nacv'] = metrics['nacv']
                all_partners[key][partner_name]['influence_count'] = metrics['count']
            else:
                all_partners[key][partner_name] = metrics
    
    for key, partners in service.items():
        for partner_name, metrics in partners.items():
            if partner_name in all_partners[key]:
                all_partners[key][partner_name]['service_nacv'] = metrics['nacv']
                all_partners[key][partner_name]['service_count'] = metrics['count']
            else:
                all_partners[key][partner_name] = metrics
    
    # Compute summary metrics
    dashboard_data = {
        'generated_at': datetime.now().isoformat(),
        'total_records': len(df),
        'partners': {}
    }
    
    # Flatten for JSON
    for (geo, territory, product, source, deal_type, month), partners in all_partners.items():
        filter_key = f"{geo}|{territory}|{product}|{source}|{deal_type}|{month}"
        
        dashboard_data['partners'][filter_key] = {
            'filters': {
                'geo': geo,
                'territory': territory,
                'product': product,
                'source': source,
                'deal_type': deal_type,
                'month': month
            },
            'partners': {}
        }
        
        for partner_name, metrics in partners.items():
            nacv = metrics.get('nacv', 0)
            count = metrics.get('count', 0)
            won = metrics.get('won', 0)
            win_rate = (won / count * 100) if count > 0 else 0
            avg_size = nacv / count if count > 0 else 0
            
            dashboard_data['partners'][filter_key]['partners'][partner_name] = {
                'name': partner_name,
                'tier': metrics.get('tier', 'Unclassified'),
                'top_level': metrics.get('top_level', partner_name),
                'contribution_type': metrics.get('contribution_type', 'Unknown'),
                'nacv': round(nacv, 2),
                'count': count,
                'won': won,
                'lost': metrics.get('lost', 0),
                'win_rate': round(win_rate, 1),
                'avg_deal_size': round(avg_size, 2)
            }
    
    return dashboard_data

# ============================================================================
# MAIN
# ============================================================================

if __name__ == '__main__':
    print("=" * 80)
    print("PARTNER PERFORMANCE CUBE - DATA AGGREGATION")
    print("=" * 80)
    
    # Load lookups
    print("\n1. Loading partner and team lookups...")
    partner_lookup = load_partner_lookups()
    team_lookup = load_team_lookups()
    print(f"   - {len(partner_lookup)} partners loaded")
    print(f"   - {len(team_lookup)} team mappings loaded")
    
    # Load and clean data
    print("\n2. Loading and cleaning SQL export...")
    df = load_data('/mnt/user-data/uploads/JA_SQL_EXPORT.xlsx')
    
    # Enrich
    print("\n3. Enriching with partner classifications and team mappings...")
    df = enrich_data(df, partner_lookup, team_lookup)
    
    # Aggregate
    print("\n4. Pre-computing all aggregations...")
    dashboard_data = build_dashboard_data(df)
    
    # Save
    output_path = '/home/claude/dashboard_data.json'
    with open(output_path, 'w') as f:
        json.dump(dashboard_data, f, indent=2)
    
    print(f"\n5. Output saved to {output_path}")
    print(f"   - {len(dashboard_data['partners'])} filter combinations")
    print(f"   - Ready for embedding in HTML dashboard")
    
    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)
